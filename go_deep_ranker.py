"""
go_deep_ranker.py — Brook's Go Deep attendee ranking module.

Given an attendee list (registration PDF or sign-in sheet), rank orgs by
priority for BD outreach:

  Tier 1: Prime-capable + existing Zoho account
  Tier 2: Prime-capable + prior/warm relationship
  Tier 3: Prime-capable + cold
  Tier 4: Engineer/consultant prime + relationship
  Tier 5: Engineer/consultant prime + cold
  Tier 6: Sub / specialty / supplier / not directly useful

Ranking model derived from Brook's transcript walking through Key Bridge
attendees ("Halmar's a big GC, Skanska's a big GC, that's a metal
fabricator"). Primary key is capability to prime the project; secondary is
Zoho Accounts relationship; tertiary is sector fit.

Sector-aware: known-primes lists differ per sector. This build targets
heavy_civil (Key Bridge, SITES conveyance). Data-center / pharma variants
plug into the same schema.
"""

import re
import pdfplumber
from dataclasses import dataclass, field, asdict
from rapidfuzz import fuzz


# ============================================================
# Sector-conditional known primes
# ============================================================

# (canonical_name, [aliases]) — canonical is the label Excel groups under.
HEAVY_CIVIL_GC_PRIMES = [
    ("Kiewit", ["Kiewit", "Kiewit Infrastructure"]),
    ("Skanska", ["Skanska", "Skanska USA", "Skanska Civil", "Skanska Koch"]),
    ("FlatironDragados", ["FlatironDragados", "Flatiron Dragados", "Flatiron", "Dragados"]),
    ("Traylor Brothers", ["Traylor Brothers", "Traylor Bros", "Traylor"]),
    ("Halmar International", ["Halmar International", "Halmar"]),
    ("S&B USA / Fay", ["Shikun Benui", "S&B USA Construction", "Shikun USA", "S&B USA", "Fay S&B", "Fay - S&B USA"]),
    ("Lane Construction", ["Lane Construction", "The Lane Construction Corporation"]),
    ("OHLA", ["OHLA", "OHLA USA", "OHL", "OHL USA"]),
    ("Walsh / Archer Western", ["Archer Western", "Walsh Group", "The Walsh Group", "Walsh Construction"]),
    ("American Bridge Company", ["American Bridge Company", "American Bridge"]),
    ("Bernards", ["Bernards"]),
    ("PCL", ["PCL", "PCL Construction", "PCL Civil Constructors"]),
    ("AECON", ["AECON", "Aecon", "Aecon Pacific Northwest"]),
    ("Fluor", ["Fluor"]),
    ("Bechtel", ["Bechtel"]),
    ("Granite Construction", ["Granite Construction", "Granite"]),
    ("Barnard Construction", ["Barnard Construction", "Barnard"]),
    ("Ames Construction", ["Ames Construction", "Ames"]),
    ("Balfour Beatty", ["Balfour Beatty"]),
    ("Manson Construction", ["Manson Construction", "Manson"]),
    ("Sundt", ["Sundt"]),
    ("Mortenson", ["Mortenson"]),
    ("AECOM Hunt", ["AECOM Hunt"]),
    ("Tutor Perini", ["Tutor Perini", "Perini"]),
    ("Ferrovial", ["Ferrovial", "Ferrovial Construction"]),
    ("Modern Continental", ["Modern Continental"]),
    ("FCC", ["FCC Construccion", "FCC"]),
    ("Webuild", ["Salini Impregilo", "Webuild"]),
    ("Sukut", ["Sukut"]),
    ("Aldridge Electric", ["Aldridge Electric", "Aldridge"]),
    ("Kokosing", ["Kokosing"]),
    ("Certerra RMA", ["Certerra RMA", "Certerra"]),
]

# Engineering / design primes — capable of leading design side, not construction
HEAVY_CIVIL_DESIGN_PRIMES = [
    ("AECOM", ["AECOM"]),  # AECOM ex-Hunt is engineering-led
    ("WSP", ["WSP"]),
    ("Stantec", ["Stantec"]),
    ("HDR", ["HDR"]),
    ("HNTB", ["HNTB"]),
    ("STV", ["STV"]),
    ("T.Y. Lin", ["T.Y. Lin", "TYLin"]),
    ("Parsons", ["Parsons"]),
    ("Michael Baker", ["Michael Baker"]),
    ("Louis Berger", ["Louis Berger"]),
    ("COWI", ["COWI", "COWI North America"]),
    ("Kimley-Horn", ["Kimley-Horn"]),
    ("Jacobs", ["Jacobs", "Jacobs Engineering"]),
    ("Arup", ["Arup"]),
    ("Mott MacDonald", ["Mott MacDonald"]),
    ("Ramboll", ["Ramboll"]),
    ("Atkins", ["Atkins"]),
    ("Kleinfelder", ["Kleinfelder", "Century Engineering a Kleinfelder"]),
    ("Brown & Caldwell", ["Brown and Caldwell", "Brown & Caldwell"]),
    ("Brierley Associates", ["Brierley Associates"]),
    ("Frank A. Olsen", ["Frank A. Olsen", "Frank A Olsen"]),
]


# ============================================================
# Zoho Accounts — live pull
# ============================================================

import os
import requests


def get_zoho_access_token() -> str | None:
    """Exchange refresh token for access token. Same pattern as crm_crossref.py."""
    client_id = os.environ.get("ZOHO_CLIENT_ID")
    client_secret = os.environ.get("ZOHO_CLIENT_SECRET")
    refresh_token = os.environ.get("ZOHO_REFRESH_TOKEN")

    if not all([client_id, client_secret, refresh_token]):
        print("Zoho credentials not set -- falling back to stub")
        return None

    try:
        r = requests.post(
            "https://accounts.zoho.com/oauth/v2/token",
            params={
                "grant_type": "refresh_token",
                "client_id": client_id,
                "client_secret": client_secret,
                "refresh_token": refresh_token,
            },
            timeout=10,
        )
        data = r.json()
        return data.get("access_token")
    except Exception as e:
        print(f"Zoho token refresh failed: {e}")
        return None


def fetch_zoho_accounts(token: str) -> dict[str, str]:
    """Pull all Account names from Zoho CRM. Returns dict of name -> 'Existing Account'."""
    accounts: dict[str, str] = {}
    page = 1
    per_page = 200

    while True:
        try:
            r = requests.get(
                "https://www.zohoapis.com/crm/v2/Accounts",
                headers={"Authorization": f"Zoho-oauthtoken {token}"},
                params={"page": page, "per_page": per_page, "fields": "Account_Name"},
                timeout=15,
            )
            data = r.json()
            records = data.get("data", [])
            if not records:
                break

            for rec in records:
                name = rec.get("Account_Name", "").strip()
                if name:
                    accounts[name] = "Existing Account"

            info = data.get("info", {})
            if not info.get("more_records", False):
                break
            page += 1

        except Exception as e:
            print(f"Zoho Accounts fetch error on page {page}: {e}")
            break

    print(f"Fetched {len(accounts)} Zoho Accounts")
    return accounts


def get_zoho_accounts() -> dict[str, str] | None:
    """Get all Zoho Accounts. Returns None if credentials missing or fetch fails."""
    token = get_zoho_access_token()
    if not token:
        return None
    accounts = fetch_zoho_accounts(token)
    if not accounts:
        return None
    return accounts


# ============================================================
# Zoho Accounts STUB — fallback when env vars not set
# ============================================================

# Stub keyed by canonical prime name (matches HEAVY_CIVIL_GC_PRIMES first tuple element).
STELIC_ACCOUNTS_STUB = {
    "Halmar International": "Prior Work",
    "Skanska": "Existing Client",
    "S&B USA / Fay": "Existing Client",
    "Lane Construction": "Prior Work",
    "OHLA": "Prior Work",
    "Aldridge Electric": "Warm Contact",
    "Walsh / Archer Western": "Prior Work",
}


# ============================================================
# Attendee model
# ============================================================

@dataclass
class Attendee:
    first_name: str = ""
    last_name: str = ""
    email: str = ""
    phone: str = ""
    organization: str = ""
    job_title: str = ""
    website: str = ""
    hq: str = ""
    offices: str = ""
    work_construction: str = ""
    work_engineering: str = ""
    work_supplier: str = ""
    work_other: str = ""
    teaming_interest: str = ""
    packages: str = ""
    country: str = ""

    # Derived
    org_normalized: str = ""
    canonical_org: str = ""    # Canonical name for grouping (Skanska variants -> "Skanska")
    role_class: str = ""       # GC_PRIME | DESIGN_PRIME | SUB | SUPPLIER | UNKNOWN
    self_declared_prime: bool = False
    zoho_status: str = "Cold"  # Existing Client | Prior Work | Warm Contact | Cold
    matched_zoho_account: str = ""
    tier: int = 6
    tier_label: str = ""
    score: int = 0
    reasoning: str = ""


# ============================================================
# Normalization
# ============================================================

_SUFFIX_RX = re.compile(
    r"\b(inc\.?|llc|l\.l\.c\.?|corp\.?|corporation|co\.?|company|"
    r"ltd\.?|limited|group|holdings|usa|us|na|america|americas|"
    r"construction|constructors|contracting|contractors|"
    r"consulting|consultants|associates|engineers?|engineering|"
    r"services|solutions)\b",
    re.IGNORECASE,
)


def normalize_org(name: str) -> str:
    if not name:
        return ""
    s = name.lower()
    s = re.sub(r"[,\.]", " ", s)
    s = re.sub(r"[^\w\s&\-]", " ", s)
    s = _SUFFIX_RX.sub(" ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def fuzzy_match(a: str, b: str) -> int:
    return fuzz.token_set_ratio(normalize_org(a), normalize_org(b))


def _first_token(name: str) -> str:
    """First substantive token of a company name, lowercased. 'The Walsh Group' -> 'walsh'."""
    stopwords = {"the", "a", "an"}
    tokens = re.findall(r"[A-Za-z0-9&]+", name.lower())
    for t in tokens:
        if t not in stopwords:
            return t
    return ""


def match_canonical(org: str, primes: list[tuple[str, list[str]]], threshold: int = 85) -> str | None:
    """Return canonical prime name if org fuzzy-matches any alias, else None.

    Enforces that the first substantive token of the alias appears in the org
    name (as a substring). Prevents 'Kiewit Infrastructure' from matching
    'Infrastructure Consulting and Engineering' via 'infrastructure' overlap.
    """
    org_lower = org.lower()
    best_canonical, best_score = None, 0
    for canonical, aliases in primes:
        for alias in aliases:
            first = _first_token(alias)
            # Must contain the primary identifying token
            if first and first not in org_lower:
                # Special-case single-letter tokens (S&B) — fall back to fuzzy only
                if len(first) > 1:
                    continue
            s = fuzzy_match(org, alias)
            if s > best_score:
                best_score, best_canonical = s, canonical
    return best_canonical if best_score >= threshold else None


def match_zoho(canonical_or_org: str, accounts: dict[str, str], threshold: int = 88) -> tuple[str, str]:
    """Return (status, matched_account_name) or ('Cold', '')."""
    # Direct hit on canonical name is exact match — no fuzzing needed.
    if canonical_or_org in accounts:
        return accounts[canonical_or_org], canonical_or_org
    best_name, best_score = "", 0
    for account in accounts:
        s = fuzzy_match(canonical_or_org, account)
        if s > best_score:
            best_score, best_name = s, account
    if best_score >= threshold:
        return accounts[best_name], best_name
    return "Cold", ""


# ============================================================
# PDF extraction — Key Bridge registration PDF shape
# ============================================================

REGISTRATION_HEADER_MARKERS = {"First Name", "Last Name", "Email", "Organization"}


def extract_registration_pdf(path: str) -> list[Attendee]:
    """Extract structured attendees from a registration-style PDF
    (Key Bridge Industry Forum shape: repeating table across pages)."""
    attendees: list[Attendee] = []

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table or len(table) < 2:
                    continue
                header = table[0]
                # Skip if header row doesn't look like a registration header
                if not any(m in " ".join(str(c) for c in header if c) for m in REGISTRATION_HEADER_MARKERS):
                    continue

                for row in table[1:]:
                    if not row or all(not c for c in row):
                        continue
                    # Pad / trim to 16 fields
                    row = list(row) + [""] * (16 - len(row))
                    row = [(c or "").replace("\n", " ").strip() for c in row]

                    a = Attendee(
                        first_name=row[0],
                        last_name=row[1],
                        email=row[2],
                        phone=row[3],
                        organization=row[4],
                        job_title=row[5],
                        website=row[6],
                        hq=row[7],
                        offices=row[8],
                        work_construction=row[9],
                        work_engineering=row[10],
                        work_supplier=row[11],
                        work_other=row[12],
                        teaming_interest=row[13],
                        packages=row[14],
                        country=row[15],
                    )
                    if not a.organization and not a.email:
                        continue
                    attendees.append(a)

    return attendees


# ============================================================
# Classification + tier assignment
# ============================================================

def classify(
    a: Attendee,
    gc_primes: list[tuple[str, list[str]]],
    design_primes: list[tuple[str, list[str]]],
) -> None:
    """Set role_class, canonical_org, self_declared_prime on attendee."""
    a.org_normalized = normalize_org(a.organization)

    ti = (a.teaming_interest or "").lower()
    a.self_declared_prime = "prime" in ti

    gc_canonical = match_canonical(a.organization, gc_primes)
    design_canonical = match_canonical(a.organization, design_primes)

    # GC prime match wins over design match (Jacobs, AECOM overlap)
    if gc_canonical:
        a.role_class = "GC_PRIME"
        a.canonical_org = gc_canonical
        return
    if design_canonical:
        a.role_class = "DESIGN_PRIME"
        a.canonical_org = design_canonical
        return

    # Not in known lists — fall back to teaming interest + work type
    wc = (a.work_construction or "").lower()
    ws = (a.work_supplier or "").lower()
    we = (a.work_engineering or "").lower()

    a.canonical_org = a.organization  # ungrouped orgs stay as-is

    if a.self_declared_prime and ("bridge" in wc or "deep foundations" in wc or "embankment" in wc):
        a.role_class = "POSSIBLE_GC_PRIME"
        return
    if ws and not wc:
        a.role_class = "SUPPLIER"
        return
    if we and not wc:
        a.role_class = "DESIGN_CONSULTANT"
        return
    if wc:
        a.role_class = "SUB"
        return
    a.role_class = "UNKNOWN"


def assign_tier(a: Attendee) -> None:
    """Assign tier (1 = highest priority) and human-readable label."""

    # Tier scoring baseline
    if a.role_class == "GC_PRIME":
        if a.zoho_status not in ("Cold", ""):
            a.tier, a.tier_label = 1, "Tier 1 — Prime GC + Relationship"
        elif a.zoho_status == "Warm Contact":
            a.tier, a.tier_label = 2, "Tier 2 — Prime GC + Warm Contact"
        else:
            a.tier, a.tier_label = 3, "Tier 3 — Prime GC (Cold)"
    elif a.role_class == "POSSIBLE_GC_PRIME":
        # Same as Tier 3 but mark reasoning
        if a.zoho_status != "Cold":
            a.tier, a.tier_label = 3, "Tier 3 — Possible Prime GC + Relationship"
        else:
            a.tier, a.tier_label = 4, "Tier 4 — Possible Prime GC (Cold)"
    elif a.role_class == "DESIGN_PRIME":
        if a.zoho_status != "Cold":
            a.tier, a.tier_label = 4, "Tier 4 — Design Prime + Relationship"
        else:
            a.tier, a.tier_label = 5, "Tier 5 — Design Prime (Cold)"
    elif a.role_class in ("DESIGN_CONSULTANT", "SUB"):
        a.tier, a.tier_label = 6, "Tier 6 — Sub / Consultant"
    elif a.role_class == "SUPPLIER":
        a.tier, a.tier_label = 7, "Tier 7 — Supplier / Specialty"
    else:
        a.tier, a.tier_label = 8, "Tier 8 — Unclassified"


def build_reasoning(a: Attendee) -> str:
    bits = []
    bits.append(f"class={a.role_class}")
    if a.zoho_status != "Cold":
        bits.append(f"zoho={a.zoho_status}")
    if a.self_declared_prime:
        bits.append("self-declared prime")
    return "; ".join(bits)


# ============================================================
# Top-level pipeline
# ============================================================

def go_deep(
    path: str,
    sector: str = "heavy_civil",
    accounts: dict[str, str] | None = None,
) -> list[Attendee]:
    if sector == "heavy_civil":
        gc_primes = HEAVY_CIVIL_GC_PRIMES
        design_primes = HEAVY_CIVIL_DESIGN_PRIMES
    else:
        # Default to heavy civil primes — Zoho relationship matching
        # still works regardless of sector
        gc_primes = HEAVY_CIVIL_GC_PRIMES
        design_primes = HEAVY_CIVIL_DESIGN_PRIMES

    if accounts is None:
        # Try live Zoho pull first, fall back to stub
        live = get_zoho_accounts()
        accounts = live if live else STELIC_ACCOUNTS_STUB

    attendees = extract_registration_pdf(path)
    # Drop rows with no org
    attendees = [a for a in attendees if (a.organization or "").strip()]
    print(f"Extracted {len(attendees)} attendee rows")

    for a in attendees:
        classify(a, gc_primes, design_primes)
        # Match Zoho by canonical (falls back to org name for un-canonicalized)
        lookup_name = a.canonical_org or a.organization
        a.zoho_status, a.matched_zoho_account = match_zoho(lookup_name, accounts)
        a.reasoning = build_reasoning(a)
        assign_tier(a)

    attendees.sort(key=lambda x: (x.tier, x.canonical_org.lower(), x.organization.lower()))
    return attendees


# ============================================================
# Excel output
# ============================================================

def write_excel(attendees: list[Attendee], out_path: str) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # Sheet 1: Ranked orgs (one row per org, top contact)
    ws = wb.active
    ws.title = "Ranked Orgs"

    org_map: dict[str, list[Attendee]] = {}
    for a in attendees:
        # Prefer canonical for grouping — Skanska Koch + Skanska Civil US -> "Skanska"
        key = a.canonical_org.lower().strip() if a.canonical_org else (a.org_normalized or a.organization.lower().strip())
        org_map.setdefault(key, []).append(a)

    # Sort by best (lowest) tier per org
    org_rows = []
    for key, rows in org_map.items():
        rows.sort(key=lambda x: x.tier)
        best = rows[0]
        org_rows.append((best, rows))
    org_rows.sort(key=lambda x: (x[0].tier, x[0].organization.lower()))

    ws.append([
        "Tier", "Tier Label", "Organization (Canonical)", "Registered As",
        "Zoho Status", "Role Class", "Self-Declared Prime",
        "Top Contact", "Title", "Email", "Phone",
        "Website", "HQ", "Contract Packages",
        "Reasoning", "# Reps",
    ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    tier_colors = {
        1: "C6EFCE", 2: "E2EFDA", 3: "FFEB9C",
        4: "FFF2CC", 5: "FCE4D6", 6: "F2F2F2",
        7: "F2F2F2", 8: "F2F2F2",
    }

    for best, rows in org_rows:
        row_num = ws.max_row + 1
        # Show distinct registered-as names in the group
        registered_variants = sorted({r.organization for r in rows if r.organization})
        registered_as = " | ".join(registered_variants[:3])
        if len(registered_variants) > 3:
            registered_as += f" (+{len(registered_variants) - 3} more)"
        ws.append([
            best.tier,
            best.tier_label,
            best.canonical_org or best.organization,
            registered_as,
            best.zoho_status,
            best.role_class,
            "Yes" if best.self_declared_prime else "",
            f"{best.first_name} {best.last_name}".strip(),
            best.job_title,
            best.email,
            best.phone,
            best.website,
            best.hq,
            best.packages[:150],
            best.reasoning,
            len(rows),
        ])
        color = tier_colors.get(best.tier, "FFFFFF")
        for cell in ws[row_num]:
            cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Column widths (16 columns now)
    widths = [6, 40, 28, 40, 16, 20, 8, 24, 28, 32, 16, 26, 22, 36, 36, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Sheet 2: All attendees (flat)
    ws2 = wb.create_sheet("All Attendees")
    ws2.append([
        "Tier", "Organization", "First Name", "Last Name", "Job Title",
        "Email", "Phone", "Role Class", "Zoho Status", "Teaming Interest",
    ])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
    for a in attendees:
        ws2.append([
            a.tier, a.organization, a.first_name, a.last_name, a.job_title,
            a.email, a.phone, a.role_class, a.zoho_status, a.teaming_interest[:100],
        ])
    for col_letter, w in zip("ABCDEFGHIJ", [6, 40, 14, 16, 30, 32, 16, 20, 16, 40]):
        ws2.column_dimensions[col_letter].width = w
    ws2.freeze_panes = "A2"

    # Sheet 3: Tier summary
    ws3 = wb.create_sheet("Tier Summary")
    ws3.append(["Tier", "Label", "# Orgs", "# People"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    tier_counts: dict[int, dict] = {}
    for best, rows in org_rows:
        d = tier_counts.setdefault(best.tier, {"label": best.tier_label, "orgs": 0, "people": 0})
        d["orgs"] += 1
        d["people"] += len(rows)
    for tier in sorted(tier_counts):
        d = tier_counts[tier]
        ws3.append([tier, d["label"], d["orgs"], d["people"]])
    for col_letter, w in zip("ABCD", [6, 44, 10, 10]):
        ws3.column_dimensions[col_letter].width = w
    ws3.freeze_panes = "A2"

    wb.save(out_path)


# ============================================================
# Apollo contact enrichment
# ============================================================

APOLLO_TARGET_TITLES = [
    "preconstruction", "pre-construction", "estimating", "estimator",
    "business development", "project executive", "vice president",
    "director", "chief estimator", "proposal", "pursuit",
]

APOLLO_MAX_ORGS = 15  # Cap orgs to search per run
APOLLO_MAX_PER_ORG = 3  # Max contacts per org


def apollo_enrich(org_names: list[str]) -> dict[str, list[dict]]:
    """Search Apollo for decision-maker contacts at each org.
    Two-step: search for IDs, then bulk_match to reveal details.
    Returns {org_name: [{name, title, email, linkedin}]}."""
    api_key = os.environ.get("APOLLO_API_KEY")
    if not api_key:
        print("APOLLO_API_KEY not set -- skipping enrichment")
        return {}

    headers = {"X-Api-Key": api_key, "Content-Type": "application/json"}
    results: dict[str, list[dict]] = {}

    for org in org_names[:APOLLO_MAX_ORGS]:
        try:
            # Step 1: Search for people IDs
            r = requests.post(
                "https://api.apollo.io/api/v1/mixed_people/api_search",
                headers=headers,
                json={
                    "q_organization_name": org,
                    "person_titles": APOLLO_TARGET_TITLES,
                    "person_seniorities": ["vp", "director", "c_suite", "owner", "partner"],
                    "per_page": APOLLO_MAX_PER_ORG,
                    "page": 1,
                },
                timeout=15,
            )

            if r.status_code != 200:
                print(f"Apollo search failed for {org}: {r.status_code}")
                continue

            people = r.json().get("people", [])
            if not people:
                continue

            # Step 2: Reveal contacts via bulk_match
            person_ids = [p["id"] for p in people if p.get("id")]
            if not person_ids:
                continue

            r2 = requests.post(
                "https://api.apollo.io/api/v1/people/bulk_match",
                headers=headers,
                json={
                    "details": [{"id": pid} for pid in person_ids],
                    "reveal_personal_emails": True,
                },
                timeout=15,
            )

            if r2.status_code != 200:
                print(f"Apollo reveal failed for {org}: {r2.status_code}")
                continue

            matches = r2.json().get("matches", [])
            contacts = []
            for p in matches:
                if not p:
                    continue
                contacts.append({
                    "name": f"{p.get('first_name', '')} {p.get('last_name', '')}".strip(),
                    "title": p.get("title", ""),
                    "email": p.get("email", ""),
                    "phone": "",
                    "linkedin": p.get("linkedin_url", ""),
                    "city": p.get("city", ""),
                    "state": p.get("state", ""),
                })

            if contacts:
                results[org] = contacts
                print(f"Apollo: {org} -> {len(contacts)} contacts")

        except Exception as e:
            print(f"Apollo error for {org}: {e}")

    print(f"Apollo enrichment: {len(results)} orgs, {sum(len(v) for v in results.values())} contacts")
    return results


# ============================================================
# JSON output (for API responses)
# ============================================================

def to_json(attendees: list[Attendee], project_name: str = "", apollo_contacts: dict[str, list[dict]] | None = None) -> dict:
    """Build the JSON response for the frontend."""

    # Group by canonical org
    org_map: dict[str, list[Attendee]] = {}
    for a in attendees:
        key = a.canonical_org.lower().strip() if a.canonical_org else (
            a.org_normalized or a.organization.lower().strip()
        )
        org_map.setdefault(key, []).append(a)

    org_rows = []
    for key, rows in org_map.items():
        rows.sort(key=lambda x: x.tier)
        best = rows[0]
        org_rows.append((best, rows))
    org_rows.sort(key=lambda x: (x[0].tier, x[0].organization.lower()))

    # Tier summary
    tier_counts: dict[int, dict] = {}
    for best, rows in org_rows:
        d = tier_counts.setdefault(best.tier, {
            "label": best.tier_label, "orgs": 0, "people": 0
        })
        d["orgs"] += 1
        d["people"] += len(rows)

    tiers = []
    for tier in sorted(tier_counts):
        d = tier_counts[tier]
        tiers.append({
            "tier": tier,
            "label": d["label"],
            "orgs": d["orgs"],
            "people": d["people"],
        })

    # Top orgs (Tier 1-3 for the ranked table)
    top_orgs = []
    for best, rows in org_rows:
        if best.tier > 3:
            continue
        canonical = best.canonical_org or best.organization
        org_entry = {
            "tier": best.tier,
            "canonical": canonical,
            "zoho": best.zoho_status,
            "reps": len(rows),
            "contact": f"{best.first_name} {best.last_name}".strip(),
            "email": best.email,
            "title": best.job_title,
        }
        # Attach Apollo contacts if available
        if apollo_contacts:
            # Try matching by canonical name or org name
            ac = apollo_contacts.get(canonical, [])
            if not ac:
                ac = apollo_contacts.get(best.organization, [])
            if ac:
                org_entry["apolloContacts"] = ac
        top_orgs.append(org_entry)

    return {
        "project": project_name,
        "totalAttendees": len(attendees),
        "totalOrgs": len(org_rows),
        "tiers": tiers,
        "topOrgs": top_orgs,
    }


# ============================================================
# CLI entry point
# ============================================================

if __name__ == "__main__":
    import sys
    import json

    path = sys.argv[1] if len(sys.argv) > 1 else "keybridge.pdf"
    sector = sys.argv[2] if len(sys.argv) > 2 else "heavy_civil"
    project = sys.argv[3] if len(sys.argv) > 3 else "Untitled Project"
    out_excel = sys.argv[4] if len(sys.argv) > 4 else "go_deep_output.xlsx"

    attendees = go_deep(path, sector=sector)
    write_excel(attendees, out_excel)

    result = to_json(attendees, project_name=project)
    print(json.dumps(result))